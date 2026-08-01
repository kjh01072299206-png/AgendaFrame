"""Build a human-annotation packet from a researcher-provided article export.

The packet deliberately contains metadata and non-reversible body hashes only.
It never writes article bodies, HTML, or model labels. A case is a candidate
cluster for two independent annotators to verify; it is not a pre-filled gold
label.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

HEADER_ALIASES = {
    "article_id": ("뉴스 식별자", "article_id", "id"),
    "date": ("일자", "date", "published_at"),
    "outlet": ("언론사", "outlet", "source"),
    "title": ("제목", "title", "headline"),
    "body": ("본문", "body", "text"),
    "url": ("URL", "url", "link"),
    "body_status": ("본문 수집 상태", "body_status"),
}


def _find_column(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    normalized = {str(value or "").strip().lower(): index for index, value in enumerate(headers)}
    for alias in aliases:
        if alias.lower() in normalized:
            return normalized[alias.lower()]
    return None


def _safe_text(value: object) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def _date_text(value: object) -> str | None:
    text = _safe_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{8}", text):
        return f"{text[:4]}-{text[4:6]}-{text[6:]}"
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).isoformat() if value.tzinfo else value.isoformat()
    return text


def _body_hash(body: str) -> str | None:
    return hashlib.sha256(body.encode("utf-8")).hexdigest() if body else None


def _title_key(title: str) -> str:
    # Exact/near-exact headline groups are only candidate clusters. Humans
    # still decide whether they describe the same event.
    value = re.sub(r"[^0-9A-Za-z가-힣 ]+", " ", title.lower())
    tokens = [token for token in value.split() if len(token) >= 2]
    # A short leading-token key finds ordinary headline variants such as
    # “패싸움 2명 사상” versus “집단 패싸움 2명 사상”. It remains a
    # candidate key only; coders must verify the event from the linked article.
    return " ".join(tokens[:3])


def read_rows(path: Path, date_filter: str | None) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [_safe_text(value) for value in next(rows)]
    except StopIteration as error:
        raise ValueError("The workbook has no header row.") from error
    columns = {name: _find_column(headers, aliases) for name, aliases in HEADER_ALIASES.items()}
    missing = [
        name for name in ("article_id", "outlet", "title", "url", "body") if columns[name] is None
    ]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    result: list[dict[str, object]] = []
    for row in rows:
        if not row:
            continue

        def value(name: str) -> object:
            return (
                row[columns[name]]
                if columns[name] is not None and columns[name] < len(row)
                else None
            )

        date = _date_text(value("date"))
        if date_filter and not str(date or "").startswith(date_filter):
            continue
        article_id = _safe_text(value("article_id"))
        title = _safe_text(value("title"))
        url = _safe_text(value("url"))
        body = _safe_text(value("body"))
        if not article_id or not title or not url or not re.match(r"^https?://", url):
            continue
        result.append(
            {
                "article_id": article_id,
                "outlet": _safe_text(value("outlet")) or "미상",
                "title": title,
                "url": url,
                "published_at": date,
                "body_sha256": _body_hash(body),
                "body_available": bool(body),
                "body_status": _safe_text(value("body_status")) or None,
                "candidate_key": _title_key(title),
            }
        )
    workbook.close()
    return result


def build_packet(
    rows: list[dict[str, object]], dataset_version: str, max_articles: int
) -> list[dict[str, object]]:
    groups: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        if row["candidate_key"]:
            groups[str(row["candidate_key"])].append(row)

    candidates = [
        group for group in groups.values() if len({str(row["outlet"]) for row in group}) >= 2
    ]
    candidates.sort(
        key=lambda group: (
            -len({str(row["outlet"]) for row in group}),
            -len(group),
            min(str(row["article_id"]) for row in group),
        )
    )

    selected: list[dict[str, object]] = []
    used: set[str] = set()
    case_number = 0
    for group in candidates:
        group = sorted(group, key=lambda row: (str(row["outlet"]), str(row["article_id"])))
        remaining = max_articles - len(selected)
        if remaining <= 0:
            break
        group = group[:remaining]
        if len(group) < 2:
            continue
        case_number += 1
        case_id = f"real-{case_number:04d}"
        for row in group:
            article_id = str(row["article_id"])
            used.add(article_id)
            selected.append(
                {
                    "schema_version": 1,
                    "dataset_version": dataset_version,
                    "case_id": case_id,
                    "split": "locked_holdout",
                    "source": {
                        "kind": "real",
                        "provider": "researcher-provided-news-export",
                        "license_basis": "researcher-provided-workbook_pending_publisher_confirmation",
                        "rights_status": "pending_confirmation",
                    },
                    "article": {
                        key: row[key]
                        for key in (
                            "article_id",
                            "outlet",
                            "title",
                            "url",
                            "published_at",
                            "body_sha256",
                            "body_available",
                            "body_status",
                        )
                    },
                    "annotation": {
                        "status": "unlabeled",
                        "independent_annotators_required": 2,
                        "annotator_ids": [],
                        "labels": [],
                        "adjudicated": False,
                        "agreement": None,
                    },
                    "evidence_policy": {
                        "raw_body_included": False,
                        "evidence_text_included": False,
                        "evidence_locator_required": True,
                        "evidence_hash_required": True,
                    },
                    "locked": True,
                }
            )

    if not selected:
        raise ValueError("No real multi-outlet candidate group was found for the requested date.")
    return selected


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--date", default=None, help="Optional YYYY-MM-DD filter")
    parser.add_argument("--dataset-version", default="real-2026-08-01-v1")
    parser.add_argument("--max-articles", type=int, default=60)
    args = parser.parse_args()
    if args.max_articles < 2 or args.max_articles > 500:
        raise SystemExit("--max-articles must be between 2 and 500")
    rows = read_rows(args.input, args.date)
    packet = build_packet(rows, args.dataset_version, args.max_articles)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="\n") as output:
        for record in packet:
            output.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")
    print(
        json.dumps(
            {
                "rows_read": len(rows),
                "records_written": len(packet),
                "cases": len({row["case_id"] for row in packet}),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
